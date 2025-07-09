print('loading libraries')
import pandas as pd
print('loaded pandas')
from pathlib import Path
print('loaded path')
from datetime import date
print('loaded date')
import requests
print('loaded requests')
import os
print('loaded os')
import openpyxl
print('loaded openpyxl')
import time
print('loaded time')
from playwright.sync_api import Playwright, sync_playwright, expect
print('All libraries loaded')

user = "username"
pas = "password"
ap = "https://apricot.socialsolutions.com/auth"

# this log location will need to be updated if the template changes or moves
log_loc = fr"--file extention--\Orion Expense Reports Template effective 01-01-2025 T&E.xlsx" # EDIT THIS LINE

print('loading function')
def run(playwright: Playwright) -> None:

    # make a new folder to hold the log and reciepts
    folder = f"{date.today()}_Reciepts and log"
    directory = fr"--file extention--\Participant Reimbursement Folder\FY25 Receipts\{folder}" # EDIT THIS LINE
    os.makedirs(directory,exist_ok=True)

    # open apricot and download a report for the reimbursements
    browser = playwright.chromium.launch(headless=False)
    context = browser.new_context()
    page = context.new_page()
    print('downloading ')
    page.goto(ap)
    page.get_by_label("Username").click()
    page.get_by_label("Username").fill(user)
    page.get_by_placeholder("your password").click()
    page.get_by_placeholder("your password").fill(pas)
    page.get_by_role("button", name="Log In to Apricot").click()

    
    # page.get_by_role("button", name="Log In to Apricot").click()
    page.get_by_role("button", name="Continue").click()

    # Apricot likes to change things, so this continue may or may not be used, comment out if the second question about logging out other instances is not there
    page.get_by_role("button", name="Continue").click()

    page.get_by_role("link", name="My Reports").click()
    page.get_by_role("button", name="Management Review Reports").click()
    page.locator("#report_2281_actionsMenu").get_by_role("link", name="Run").click()
    page.wait_for_load_state('networkidle')

    # download the report to excel
    page.get_by_role("button", name="Export").click()
    with page.expect_download() as download_info:
        page.get_by_role("button", name="Export").click()
    download = download_info.value
    
    # take the downloaded report and assign it to a dataframe, updated to complete 3 times for each sheet in the report
    print('parsing through report')
    # participant reimbursement
    reim_df = pd.read_excel(download.path(),sheet_name='Active Reimbursements - Rows') 
    reim_df = reim_df.fillna('')
    reim_df.columns = reim_df.columns.str.replace(" ","_")

    # miles log
    miles_df = pd.read_excel(download.path(),sheet_name='Miles Log - Rows') 
    miles_df = miles_df.fillna('')
    miles_df.columns = miles_df.columns.str.replace(" ","_")  
    
    # non participant reimbursement
    np_df = pd.read_excel(download.path(),sheet_name='Non Participant Reimbur - Rows') 
    np_df = np_df.fillna('')
    np_df.columns = np_df.columns.str.replace(" ","_") 

    reportlist = [reim_df,miles_df,np_df]      

    # drop rows that contain processed = yes, because we dont need to process them again
    print('removing processed rows')

    # use enumerate to properly loop through the dataframes, then assign the dataframe back to the original value
    for i, df in enumerate(reportlist):
        df = df[df['Processed'] != 'Yes'].copy()

        # then we need to make a url for each record ID
        df['form_loc'] = "https://apricot.socialsolutions.com/document/edit/id/" + df["Record_ID"].astype(str)
        reportlist[i] = df

    # reassign the dataframes to the updated versions from the report list, 
    reim_df, miles_df, np_df = reportlist

    # change the formatting of the dates to better handle them
    reim_df['Date_of_Request'] = reim_df['Date_of_Request'].str.replace('/','-',regex=True)
    miles_df['Date_of_Request'] = miles_df['Date_of_Request'].str.replace('/','-',regex=True)
    np_df['Date_of_Request'] = np_df['Date_of_Request'].str.replace('/','-',regex=True)

    # in staff name remove the record ID and extra whitespace
    reim_df['Staff_Name'] = (reim_df['Staff_Name'].str.replace(r'\(\d*\)','',regex=True).str.strip().str.replace(r'\s+', ' ', regex=True))
    miles_df['Staff_Name'] = (miles_df['Staff_Name'].str.replace(r'\(\d*\)','',regex=True).str.strip().str.replace(r'\s+', ' ', regex=True))
    np_df['Staff_Name'] = (np_df['Staff_Name'].str.replace(r'\(\d*\)','',regex=True).str.strip().str.replace(r'\s+', ' ', regex=True))
    
    # and reassign the list to the udated dfs
    reportlist = [reim_df,miles_df,np_df]

    print(reim_df)
    print(miles_df)
    print(np_df)

    # loop through each report to then loop through each row of each report
    for i, df in enumerate(reportlist):
        # create an empty list to append onto the df at the end
        receipts = []

        # go to each url and parse the page to find the reciept download link
        for row in df.itertuples():
            if 'Name' in df.columns:
                print(f"obtaining receipt for {row.Name}")
            else:
                print(f'processing {row.Record_ID} for {row.Staff_Name}')
            page.goto(row.form_loc)
            page.wait_for_load_state('networkidle')
            page.get_by_role("heading", name="System Fields").click()
            # if troubleshooting is needed and the script needs to rerun multiple times. comment out the line below so it doesnt trigger the processed field_____________________
            page.get_by_label("Yes", exact=True).check()
            page.get_by_label("Date Processed").click()
            page.get_by_label("Date Processed").fill(date.today().strftime('%m/%d/%Y'))
            page.get_by_role("button", name="Save Record").click()
            page.wait_for_load_state('networkidle')
            page.get_by_role("button", name="Continue").click()
            page.wait_for_load_state('networkidle')

            # download the file based on the div ID, and the a is the a href link
            if i != 1:
                try:
                    with page.expect_download() as download_info:
                        # download the reciept from the ppt reimbursement
                        if i == 0:
                            page.locator('#field_3804_link a').click()
                        # download the reciept from the non ppt reimbursement
                        if i == 2:
                            page.locator('#field_5373_link a').click()
                    download = download_info.value

                    suffix = Path(download.suggested_filename).suffix
                    if 'Name' in df.columns:
                        receipt_name = f"{row.Staff_Name}_{row.Name}_{row.Date_of_Request}_{row.Record_ID}{suffix}"
                    else:
                        receipt_name = f'{row.Staff_Name}_{row.Date_of_Request}_{row.Record_ID}{suffix}'
                    save_path = os.path.join(directory, receipt_name)
                    download.save_as(save_path)
                    receipts.append(receipt_name)
                except:
                    receipts.append("no receipt uploaded")
                    print('no receipt uploaded')
        if i != 1:
            df['receipt'] = receipts
        reportlist[i] = df

    reim_df, miles_df, np_df = reportlist
    reportlist = [reim_df,miles_df,np_df]

    # removed these list assignments as they are added above, 
    # reim_df['receipt'] = receipts
    # miles_df['receipt'] = receipts
    # np_df['receipt'] = receipts
    print('transforming data in frames')
    for i, df in enumerate(reportlist):
        # create a new column that hyperlinks to the location of the file, must use an apply lambda function to work with the dataframe format
        if i != 1:
            df['receipt_location'] = df['receipt'].apply(
                lambda x: 'no receipt uploaded' if x=='no receipt uploaded'
                else fr'=HYPERLINK("{directory}\{x}")'
            )
        # remove puncutation from staff names
        df['Staff_Name'] = df['Staff_Name'].str.replace(r'[^a-zA-Z0-9\s]','',regex=True)
        reportlist[i] = df

    reim_df, miles_df, np_df = reportlist

    print('exporting frames to excel')
    if reim_df.empty == False:
        reim_df.to_csv(fr"{directory}\__Reimbursement_log_{date.today()}.csv") 
    if miles_df.empty == False:
        miles_df.to_csv(fr"{directory}\__Miles_log_{date.today()}.csv") 
    if np_df.empty == False:
        np_df.to_csv(fr"{directory}\__Non-Participant_Reimbursement_log_{date.today()}.csv") 

    # drop columns that were purchased with orion money, this will remove rows after the log is saved
    # but before the expense report is made
    reim_df = reim_df.drop(reim_df[reim_df['Purchase_Type:'] == 'Orion Credit Card Purchased'].index)
    miles_df = miles_df.drop(miles_df[miles_df['Purchase_Type:'] == 'Orion Credit Card Purchased'].index)
    np_df = np_df.drop(np_df[np_df['Purchase_Type:'] == 'Orion Credit Card Purchased'].index)

    # remove the id number after the employee name in two of the reports
    # miles_df['Staff_Name'] = miles_df['Staff_Name'].str.replace(r'\s*\(\d+\)','').str.strip()
    # np_df['Staff_Name'] = np_df['Staff_Name'].str.replace(r'\s*\(\d+\)','').str.strip()

    reportlist = [reim_df,miles_df,np_df]
#_____________________________________________________________________________________________
######################## all reimbursements starting with ppt forms #############################################
#_____________________________________________________________________________________________    

# need to break up the df by staff to submit separate forms for each person
    if reim_df.empty==False:
        staff_dfs = {staff: group for staff, group in reim_df.groupby('Staff_Name')}
        print(staff_dfs)

        for staff,frame in staff_dfs.items():
            
            # make a new df for miles and non participant expenses to then add onto the current excel doc
            staff_miles = miles_df[miles_df['Staff_Name']==staff]
            staff_np = np_df[np_df['Staff_Name']==staff]

            # drop the staff from the other dfs so they dont get processed later
            miles_df = miles_df.drop(miles_df[miles_df['Staff_Name']==staff].index)
            np_df = np_df.drop(np_df[np_df['Staff_Name']==staff].index)

            # make a new frame with the concat from reim_df and staff_np
            frame = pd.concat([frame,staff_np],ignore_index=True,axis=0)
            frame = frame.fillna('')

            # transform the dataframe to match the expense report
            wb = openpyxl.load_workbook(log_loc)
            ws = wb['Expense Reimbursement']
            ws2 = wb['Mileage Reimbursement']
            startrow= 9
            startcol= 2

            # fill out the name and information at the top of the form
            ws2.cell(row=2,column=5).value = staff
            ws.cell(row=2,column=5).value = staff
            if 'Mukilteo' in frame.iloc[0]['Assigned_Programs']: 
                ws.cell(row=2,column=7).value = '20 - Mukilteo'
                ws2.cell(row=2,column=7).value = '20 - Mukilteo'
            else:
                ws.cell(row=2,column=7).value = '10 - Auburn'
                ws2.cell(row=2,column=7).value = '10 - Auburn'
            ws.cell(row=5,column=5).value = '800 - T&E'
            ws2.cell(row=5,column=5).value = '800 - T&E'

            # fill out the cells that contain the reimbursement information
            for _, rows in frame.iterrows():
                print(rows)
                ws.cell(row=startrow,column=startcol).value = rows['Date_of_Request']
                if 'Snohomish County - ERSS' in rows['Sub-Service_and_Funding_Source']:
                    ws.cell(row=startrow,column=startcol+1).value = 'Yes'
                    ws.cell(row=startrow,column=startcol+2).value = 'ERSS'
                ws.cell(row=startrow,column=startcol+3).value = rows['Type_of_Reimbursement_Support']+" - "+rows['Processing_Notes']
                if rows['Expense_Type'] != '':
                    ws.cell(row=startrow,column=startcol+5).value = rows['Expense_Type']
                else:
                    ws.cell(row=startrow,column=startcol+5).value = 'Client Needs'
                ws.cell(row=startrow,column=startcol+7).value = rows['Value_of_Reimbursement']

                # check the reciept required?
                # ws.cell(row=startrow,column=startcol+8).value = True
                startrow += 1

            # add in miles to the separate sheet on the excel file
            for _, rows in staff_miles.iterrows():
                ws2.cell(row=startrow+1,column=startcol).value = rows['Date_of_Request']
                if 'ERSS' in rows['Billable_miles?']:
                    ws2.cell(row=startrow+1,column=startcol+1).value = 'Yes'
                    ws2.cell(row=startrow+1,column=startcol+2).value = 'ERSS'
                ws2.cell(row=startrow+1,column=startcol+3).value = rows['Activities_completed_during_trip']
                ws2.cell(row=startrow+1,column=startcol+5).value = rows['Processing_Notes']
                ws2.cell(row=startrow+1,column=startcol+7).value = rows['Tolls']
                ws2.cell(row=startrow+1,column=startcol+9).value = rows['Miles_Driven']
                startrow += 1
                pass
            wb.save(fr"{directory}\_{staff}_expense_report_{date.today()}.xlsx")
            wb.close()

        # reenter the form into a hidden apricot form that will then trigger an email?
            page.goto('https://apricot.socialsolutions.com/document/search/form_id/190')
            page.get_by_role("button", name="New Reimbursement log").click()
            page.get_by_placeholder("MM/DD/YYYY").click()
            page.get_by_placeholder("MM/DD/YYYY").fill(date.today().strftime('%m/%d/%Y'))
            # change location as needed
            if 'Mukilteo' in frame.iloc[0]['Assigned_Programs']: 
                page.get_by_label("Mukilteo",exact=True).check() 
            else:
                page.get_by_label("Auburn",exact=True).check()
            page.get_by_label("Reimbursement Details").click()

            # enter in line by line all the details of the form
            page.get_by_label("Reimbursement Details").type(f'{staff} \n \n')
            for _, rows in frame.iterrows():
                page.get_by_label("Reimbursement Details").type(f'{rows['Name']} - {rows['Date_of_Request']} - {rows['Type_of_Reimbursement_Support']} {rows['Processing_Notes']} - {rows['Value_of_Reimbursement']}\n')
            if staff_np.empty == False:
                for _, rows in staff_np.iterrows():
                    page.get_by_label("Reimbursement Details").type(f'{rows['Staff_Name']} - {rows['Date_of_Request']} - {rows['Expense_Type']} {rows['Processing_Notes']} - {rows['Value_of_Reimbursement']}\n')
            if staff_miles.empty == False:  
                page.get_by_label("Reimbursement Details").type('Miles Entry \n')  
                for _, rows in staff_miles.iterrows():
                    page.get_by_label("Reimbursement Details").type(f'{rows['Date_of_Request']} Miles: {rows['Miles_Driven']} Tolls: {rows['Tolls']} - {rows['Activities_completed_during_trip']} - {rows['Processing_Notes']}\n')

            page.get_by_label("Expense reports in need of approval").click()
            # link to the folder location of the expense report or at least the path location
            page.get_by_label("Expense reports in need of approval").fill("Location of expense report: \n" + fr"{directory}\_{staff}_expense_report_{date.today()}.xlsx")
            page.get_by_role("button", name="Program Access").click()
            page.get_by_label("Add all available programs to").click()
            page.get_by_role("button", name="Apply").click()
            page.get_by_role("button", name="Save Record").click()
            page.wait_for_load_state('networkidle')
            page.get_by_role("button", name="View Folder").click()

#_____________________________________________________________________________________________
######################## miles log and non participant only #############################################
#_____________________________________________________________________________________________

    # do all the same as above but just with the miles report as there will be leftover people who only have miles entries
    if miles_df.empty==False:
        staff_miles_dfs = {staff: group for staff, group in miles_df.groupby('Staff_Name')}
        print(staff_miles_dfs)

        for staff,frame in staff_miles_dfs.items():
            
            # make a new df from the miles staff if they have non participant records as well
            staff_np = np_df[np_df['Staff_Name']==staff]

            # drop the staff from the other dfs so they dont get processed later
            np_df = np_df.drop(np_df[np_df['Staff_Name']==staff].index)

            # transform the dataframe to match the expense report
            wb = openpyxl.load_workbook(log_loc)
            ws = wb['Expense Reimbursement']
            ws2 = wb['Mileage Reimbursement']
            startrow= 9
            startcol= 2

            # fill out the name and information at the top of the form
            ws2.cell(row=2,column=5).value = staff
            ws.cell(row=2,column=5).value = staff
            if 'Mukilteo' in frame.iloc[0]['Assigned_Programs']: 
                ws.cell(row=2,column=7).value = '20 - Mukilteo'
                ws2.cell(row=2,column=7).value = '20 - Mukilteo'
            else:
                ws.cell(row=2,column=7).value = '10 - Auburn'
                ws2.cell(row=2,column=7).value = '10 - Auburn'
            ws.cell(row=5,column=5).value = '800 - T&E'
            ws2.cell(row=5,column=5).value = '800 - T&E'

            # fill out the cells that contain the reimbursement information
            if staff_np.empty == False:
                for _, rows in staff_np.iterrows():
                    ws.cell(row=startrow,column=startcol).value = rows['Date_of_Request']
                    ws.cell(row=startrow,column=startcol+3).value = rows['Processing_Notes']
                    if rows['Expense_Type'] != '':
                        ws.cell(row=startrow,column=startcol+5).value = rows['Expense_Type']
                    else:
                        ws.cell(row=startrow,column=startcol+5).value = 'Client Needs'
                    ws.cell(row=startrow,column=startcol+7).value = rows['Value_of_Reimbursement']

                    # check the reciept required?
                    # ws.cell(row=startrow,column=startcol+8).value = True
                    startrow += 1

            # add in miles to the separate sheet on the excel file
            for _, rows in frame.iterrows():
                ws2.cell(row=startrow+1,column=startcol).value = rows['Date_of_Request']
                if 'ERSS' in rows['Billable_miles?']:
                    ws2.cell(row=startrow+1,column=startcol+1).value = 'Yes'
                    ws2.cell(row=startrow+1,column=startcol+2).value = 'ERSS'
                ws2.cell(row=startrow+1,column=startcol+3).value = rows['Activities_completed_during_trip']
                ws2.cell(row=startrow+1,column=startcol+5).value = rows['Processing_Notes']
                ws2.cell(row=startrow+1,column=startcol+7).value = rows['Tolls']
                ws2.cell(row=startrow+1,column=startcol+9).value = rows['Miles_Driven']
                startrow += 1
                pass
            wb.save(fr"{directory}\_{staff}_expense_report_{date.today()}.xlsx")
            wb.close()

        # reenter the form into a hidden apricot form that will then trigger an email?
            page.goto('https://apricot.socialsolutions.com/document/search/form_id/190')
            page.get_by_role("button", name="New Reimbursement log").click()
            page.get_by_placeholder("MM/DD/YYYY").click()
            page.get_by_placeholder("MM/DD/YYYY").fill(date.today().strftime('%m/%d/%Y'))
            # change location as needed
            if 'Mukilteo' in frame.iloc[0]['Assigned_Programs']: 
                page.get_by_label("Mukilteo",exact=True).check() 
            else:
                page.get_by_label("Auburn",exact=True).check()
            page.get_by_label("Reimbursement Details").click()
            # enter in line by line all the details of the form
            page.get_by_label("Reimbursement Details").type(f'{staff} \n \n')
            
            if staff_np.empty == False:
                for _, rows in staff_np.iterrows():
                    page.get_by_label("Reimbursement Details").type(f'{rows['Staff_Name']} - {rows['Date_of_Request']} - {rows['Expense_Type']} {rows['Processing_Notes']} - {rows['Value_of_Reimbursement']}\n')

            for _, rows in frame.iterrows():
                page.get_by_label("Reimbursement Details").type(f'{rows['Date_of_Request']} Miles: {rows['Miles_Driven']} Tolls: {rows['Tolls']} - {rows['Activities_completed_during_trip']} - {rows['Processing_Notes']}\n')
            
            page.get_by_label("Expense reports in need of approval").click()
            # link to the folder location of the expense report or at least the path location
            page.get_by_label("Expense reports in need of approval").fill("Location of expense report: \n" + fr"{directory}\_{staff}_expense_report_{date.today()}.xlsx")
            page.get_by_role("button", name="Program Access").click()
            page.get_by_label("Add all available programs to").click()
            page.get_by_role("button", name="Apply").click()
            page.get_by_role("button", name="Save Record").click()
            page.wait_for_load_state('networkidle')
            page.get_by_role("button", name="View Folder").click()

#_____________________________________________________________________________________________
######################## Non participant only #############################################
#_____________________________________________________________________________________________

    # do all the same as above but just with the non participant report that will pull any outliers
    if np_df.empty==False:
        staff_np_dfs = {staff: group for staff, group in np_df.groupby('Staff_Name')}
        print(staff_np_dfs)

        for staff,frame in staff_np_dfs.items():

            # transform the dataframe to match the expense report
            wb = openpyxl.load_workbook(log_loc)
            ws = wb['Expense Reimbursement']
            startrow= 9
            startcol= 2

            # fill out the name and information at the top of the form
            ws2.cell(row=2,column=5).value = staff
            ws.cell(row=2,column=5).value = staff
            if 'Mukilteo' in frame.iloc[0]['Assigned_Programs']: 
                ws.cell(row=2,column=7).value = '20 - Mukilteo'
                ws2.cell(row=2,column=7).value = '20 - Mukilteo'
            else:
                ws.cell(row=2,column=7).value = '10 - Auburn'
                ws2.cell(row=2,column=7).value = '10 - Auburn'
            ws.cell(row=5,column=5).value = '800 - T&E'
            ws2.cell(row=5,column=5).value = '800 - T&E'

            # fill out the cells that contain the reimbursement information
            for _, rows in frame.iterrows():
                ws.cell(row=startrow,column=startcol).value = rows['Date_of_Request']
                ws.cell(row=startrow,column=startcol+3).value = rows['Processing_Notes']
                ws.cell(row=startrow,column=startcol+5).value = rows['Expense_Type']
                ws.cell(row=startrow,column=startcol+7).value = rows['Value_of_Reimbursement']

                # check the reciept required?
                # ws.cell(row=startrow,column=startcol+8).value = True
                startrow += 1

            wb.save(fr"{directory}\_{staff}_expense_report_{date.today()}.xlsx")
            wb.close()

        # reenter the form into a hidden apricot form that will then trigger an email?
            page.goto('https://apricot.socialsolutions.com/document/search/form_id/190')
            page.get_by_role("button", name="New Reimbursement log").click()
            page.get_by_placeholder("MM/DD/YYYY").click()
            page.get_by_placeholder("MM/DD/YYYY").fill(date.today().strftime('%m/%d/%Y'))
            # change location as needed
            if 'Mukilteo' in frame.iloc[0]['Assigned_Programs']: 
                page.get_by_label("Mukilteo",exact=True).check() 
            else:
                page.get_by_label("Auburn",exact=True).check()
            page.get_by_label("Reimbursement Details").click()
            # enter in line by line all the details of the form
            page.get_by_label("Reimbursement Details").type(f'{staff} \n \n')
            
            for _, rows in frame.iterrows():
                    page.get_by_label("Reimbursement Details").type(f'{rows['Staff_Name']} - {rows['Date_of_Request']} - {rows['Expense_Type']} {rows['Processing_Notes']} - {rows['Value_of_Reimbursement']}\n')

            page.get_by_label("Expense reports in need of approval").click()
            # link to the folder location of the expense report or at least the path location
            page.get_by_label("Expense reports in need of approval").fill("Location of expense report: \n" + fr"{directory}\_{staff}_expense_report_{date.today()}.xlsx")
            page.get_by_role("button", name="Program Access").click()
            page.get_by_label("Add all available programs to").click()
            page.get_by_role("button", name="Apply").click()
            page.get_by_role("button", name="Save Record").click()
            page.wait_for_load_state('networkidle')
            page.get_by_role("button", name="View Folder").click()


    # ---------------------
    context.close()
    browser.close()
print('function loaded')

print('run script')
with sync_playwright() as playwright:
    run(playwright)

