print('loading libraries')
import pandas as pd
print('loaded pandas')
from pathlib import Path
print('loaded path')
from datetime import date
print('loaded date')
import os
print('loaded os')
import openpyxl
print('loaded openpyxl')
from playwright.sync_api import Playwright, sync_playwright, expect
print('All libraries loaded')

user = "username"
pas = "password"
ap = "https://apricot.socialsolutions.com/auth"

# this log location will need to be updated if the template changes or moves
log_loc = fr"Redacted Path\Expense Reports Template.xlsx"

print('loading function')
def run(playwright: Playwright) -> None:

    # make a new folder to hold the log and reciepts
    folder = f"{date.today()}_Reciepts and log"
    directory = fr"Redacted Path\Participant Reimbursement Folder\Receipts\{folder}"
    os.makedirs(directory,exist_ok=True)

    # open apricot and download a report for the reimbursements
    browser = playwright.chromium.launch(headless=False)
    context = browser.new_context()
    page = context.new_page()
    print('downloading ')
    page.goto(ap)
    page.get_by_label("Username").click()
    page.get_by_label("Username").fill(user)
    page.get_by_placeholder("your password").click()
    page.get_by_placeholder("your password").fill(pas)
    page.get_by_role("button", name="Log In to Apricot").click()
    page.get_by_role("button", name="Log In to Apricot").click()
    page.get_by_role("button", name="Continue").click()
    page.get_by_role("link", name="My Reports").click()
    page.get_by_role("button", name="Management Review Reports").click()
    page.locator("#report_2281_actionsMenu").get_by_role("link", name="Run").click()
    page.wait_for_load_state('networkidle')

    # download the report to excel
    page.get_by_role("button", name="Export").click()
    with page.expect_download() as download_info:
        page.get_by_role("button", name="Export").click()
    download = download_info.value
    
    # take the downloaded report and assign it to a dataframe  
    print('parsing through report')
    reim_df = pd.read_excel(download.path())
    reim_df = reim_df.fillna('')
    reim_df.columns = reim_df.columns.str.replace(" ","_")

    # drop rows that contain processed = yes, because we dont need to process them again
    reim_df = reim_df[reim_df['Processed'] != 'Yes'].copy()

    # then we need to make a url for each record ID
    reim_df['form_loc'] = "https://apricot.socialsolutions.com/document/edit/id/" + reim_df["Record_ID"].astype(str)
    reim_df['Date_of_Request'] = reim_df['Date_of_Request'].str.replace('/','-',regex=True)
    print(reim_df)

    receipts = []

    # go to each url and parse the page to find the reciept download link
    for row in reim_df.itertuples():
        print(f"obtaining receipt for {row.Name}")
        page.goto(row.form_loc)
        page.get_by_role("heading", name="System Fields").click()
        page.get_by_label("Yes").check()
        page.get_by_role("button", name="Save Record").click()
        page.wait_for_load_state('networkidle')
        page.get_by_role("button", name="Continue").click()
        page.wait_for_load_state('networkidle')

        # download the file based on the div ID, and the a is the a href link
        try:
            with page.expect_download() as download_info:
                page.locator('#field_3804_link a').click()
            download = download_info.value

            suffix = Path(download.suggested_filename).suffix
            receipt_name = f"{row.Name}_{row.Date_of_Request}_receipt_{row.Record_ID}{suffix}"
            save_path = os.path.join(directory, receipt_name)
            download.save_as(save_path)
            receipts.append(receipt_name)
        except:
            receipts.append("no receipt uploaded")
            print('no receipt uploaded')

    reim_df['receipt'] = receipts

    # create a new column that hyperlinks to the location of the file, must use an apply lambda function to work with the dataframe format
    reim_df['receipt_location'] = reim_df['receipt'].apply(
        lambda x: 'no receipt uploaded' if x=='no receipt uploaded'
        else fr'=HYPERLINK("{directory}\{x}")'
    )
    # remove puncutation from staff names
    reim_df['Staff_Name'] = reim_df['Staff_Name'].str.replace(r'[^a-zA-Z0-9\s]','',regex=True)

    reim_df.to_csv(fr"{directory}\_Reimbursement_log_{date.today()}.csv")

    # drop columns that were purchased with company money, this will remove rows after the log is saved
    # but before the expense report is made
    reim_df = reim_df.drop(reim_df[reim_df['Purchase_Type:'] == 'Company Credit Card Purchased'].index)

    # need to break up the df by staff to submit separate forms for each person
    staff_dfs = {staff: group for staff, group in reim_df.groupby('Staff_Name')}
    print(staff_dfs)

    for staff,frame in staff_dfs.items():
        # transform the dataframe to match the expense report
        wb = openpyxl.load_workbook(log_loc)
        ws = wb['Expense Reimbursement']
        startrow= 9
        startcol= 2

        # fill out the name and information at the top of the form
        ws.cell(row=2,column=4).value = staff
        if 'Mukilteo' in frame.iloc[0]['Assigned_Programs']: 
            ws.cell(row=2,column=6).value = '20 - Mukilteo'
        else:
            ws.cell(row=2,column=6).value = '10 - Auburn'
        ws.cell(row=5,column=4).value = '800 - T&E'

        # fill out the cells that contain the reimbursement information
        for _, rows in frame.iterrows():
            ws.cell(row=startrow,column=startcol).value = rows['Date_of_Request']
            ws.cell(row=startrow,column=startcol+2).value = rows['Type_of_Reimbursement_Support']+" - "+rows['Processing_Notes']
            if rows['Expense_Type'] != '':
                ws.cell(row=startrow,column=startcol+4).value = rows['Expense_Type']
            else:
                ws.cell(row=startrow,column=startcol+4).value = 'Client Needs'
            ws.cell(row=startrow,column=startcol+6).value = rows['Value_of_Reimbursement']

            # check the reciept required?
            # ws.cell(row=startrow,column=startcol+8).value = True
            startrow += 1
        wb.save(fr"{directory}\_{staff}_expense_report_{date.today()}.xlsx")
        wb.close()

    # reenter the form into a hidden apricot form that will then trigger an email?
        page.goto('https://apricot.socialsolutions.com/document/search/form_id/redacted')
        page.get_by_role("button", name="New Reimbursement log").click()
        page.get_by_placeholder("MM/DD/YYYY").click()
        page.get_by_placeholder("MM/DD/YYYY").fill(date.today().strftime('%m/%d/%Y'))
        # change location as needed
        if 'Mukilteo' in frame.iloc[0]['Assigned_Programs']: 
            page.get_by_label("Mukilteo").check() 
        else:
            page.get_by_label("Auburn").check()
        page.get_by_label("Reimbursement Details").click()
        # enter in line by line all the details of the form
        page.get_by_label("Reimbursement Details").type(f'{staff} \n \n')
        for _, rows in frame.iterrows():
            page.get_by_label("Reimbursement Details").type(f'{rows['Name']} - {rows['Date_of_Request']} - {rows['Type_of_Reimbursement_Support']} {rows['Processing_Notes']} - {rows['Value_of_Reimbursement']}\n')
        page.get_by_label("Expense reports in need of approval").click()
        # link to the folder location of the expense report or at least the path location
        page.get_by_label("Expense reports in need of approval").fill("Location of expense report: \n" + fr"{directory}\_{staff}_expense_report_{date.today()}.xlsx")
        page.get_by_role("button", name="Save Record").click()
        page.wait_for_load_state('networkidle')
        page.get_by_role("button", name="View Folder").click()

    # mark the form as processed to get it out of the report

    # reim_df.to_csv(fr"{directory}\_Reimbursement_log_{date.today()}.csv")

    # ---------------------
    context.close()
    browser.close()
print('function loaded')

print('run script')
with sync_playwright() as playwright:
    run(playwright)